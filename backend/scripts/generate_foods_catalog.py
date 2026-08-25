"""Build data/foods_catalog.csv from dish_nutrition.csv + a local INDB xlsx.

Dev-only (needs openpyxl). Production upsert reads the CSV, not Excel.
Does not dump all INDB recipes. Curated staples only.

Place `data/Anuvaad_INDB_2024.11.xlsx` locally (gitignored). Upstream:
https://github.com/lindsayjaacks/Indian-Nutrient-Databank-INDB-/blob/main/INDB.xlsx

  python scripts/generate_foods_catalog.py
"""

from __future__ import annotations

import csv
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from services.food.services.food_classes_service import FOOD_CLASS_LABELS  # noqa: E402

DISH_CSV = ROOT / "data" / "dish_nutrition.csv"
INDB_XLSX = ROOT / "data" / "Anuvaad_INDB_2024.11.xlsx"
OUT_CSV = ROOT / "data" / "foods_catalog.csv"

QUANTIZE = Decimal("0.01")
SERVING_MIN = Decimal("10")
SERVING_MAX = Decimal("400")
MAX_FAT_PER_100G = Decimal("40")

# Extra typed-log staples (not in the 30 detect classes). Numbers come from INDB.
STAPLES: tuple[tuple[str, str, str], ...] = (
    ("ASC096", "roti", "chapati,chapatti,phulka,wheat roti"),
    ("ASC097", "paratha", "parantha,plain paratha,plain parantha"),
    ("ASC098", "aloo-paratha", "aloo parantha,potato paratha"),
    ("ASC099", "mooli-paratha", "radish paratha,mooli parantha"),
    ("ASC105", "paneer-paratha", "paneer parantha"),
    ("ASC142", "naan", "nan,tandoori naan"),
    ("ASC150", "makki-roti", "makki ki roti,corn roti"),
    ("ASC472", "paushtik-roti", "poshtik roti"),
    ("ASC114", "plain-pulao", "pulao,veg pulao,plain pulav"),
    ("ASC115", "mixed-vegetable-pulao", "vegetable pulao,mix veg pulao"),
    ("ASC122", "mutton-biryani", "gosht biryani,mutton biriyani"),
    ("ASC123", "vegetable-biryani", "veg biryani,veg biriyani"),
    ("ASC124", "lemon-rice", "chitranna,pulihora,elumichai sadam"),
    ("ASC126", "curd-rice", "dahi chawal,thayir sadam,daddojanam"),
    ("ASC127", "tamarind-rice", "puliyodharai,puli sadam"),
    ("ASC151", "moong-dal", "dhuli moong,yellow moong dal,washed moong dal"),
    ("ASC152", "urad-dal", "dhuli urad,washed urad dal"),
    ("ASC155", "mixed-dal", "mix dal,dal mix"),
    ("ASC156", "whole-moong-dal", "sabut moong,moong whole"),
    ("ASC157", "masoor-dal", "whole masoor,sabut masoor"),
    ("ASC167", "sambar", "sambhar"),
    ("ASC056", "boiled-egg", "ubla anda,hard boiled egg"),
    ("ASC057", "fried-egg", "anda fry"),
    ("ASC059", "anda-bhurji", "scrambled egg,egg bhurji,ande ki bhurji"),
    ("ASC061", "omelette", "omlet,plain omelette"),
    ("ASC240", "chicken-curry", "murgh curry,chicken masala"),
    ("ASC177", "baingan-bharta", "baingan ka bharta,brinjal bharta"),
    ("ASC190", "aloo-matar", "pea potato curry,aloo mattar"),
    ("ASC191", "matar-paneer", "pea paneer,mutter paneer"),
    ("ASC195", "paneer-curry", "paneer gravy"),
    ("ASC226", "kadhai-paneer", "kadai paneer"),
    ("ASC173", "cabbage-peas", "pattagobhi matar,cabbage and peas"),
    ("ASC273", "cucumber-raita", "kheere ka raita,kheera raita"),
    ("ASC269", "onion-tomato-raita", "tamatar pyaaz raita"),
    ("ASC146", "masala-dosa", "masala dose"),
    ("ASC147", "rava-dosa", "suji dosa,semolina dosa"),
    ("ASC474", "dhokla", "gujarati dhokla"),
    ("BFP039", "rava-upma", "suji upma,semolina upma,upma"),
    ("BFP043", "vegetable-upma", "veg upma"),
    ("BFP144", "plain-khichdi", "khichri,khitchdi,plain khichri"),
    ("BFP176", "rasam", "puli rasam,saaru,charu"),
    ("BFP177", "lemon-rasam", "nimmakaya rasam,elumichai rasam"),
    ("BFP240", "egg-curry", "anda curry"),
    ("OSR104", "methi-thepla", "thepla,fenugreek thepla"),
    ("OSR105", "rice-puttu", "ari puttu,puttu"),
    ("BFP153", "appam", "palappam,hoppers"),
    ("OSR139", "dal-makhani", "daal makhani,kali dal"),
    ("ASC022", "salted-lassi", "namkeen lassi,chaas lassi"),
    ("OSR010", "saffron-milk", "kesariya doodh,kesar doodh"),
    ("ASC047", "daliya", "cracked wheat porridge,meetha daliya"),
    ("ASC293", "suji-halwa", "semolina halwa,rava kesari"),
    ("ASC295", "gajar-halwa", "carrot halwa,gajar ka halwa"),
    ("ASC381", "paneer-tikka", "paneer shashlik,paneer tikka"),
    ("ASC385", "peanut-chutney", "mungfali chutney"),
    ("BFP439", "khaman", "khaman dhokla"),
    ("BFP576", "poshtik-khichdi", "paushtik khichdi"),
    ("OSR151", "jowar-dosa", "sorghum dosa"),
    ("BFP040", "seviyan-upma", "vermicelli upma,semiya upma"),
    ("OSR099", "sabudana-khichdi", "sago khichdi,sabudana khichri"),
    ("ASC381", "paneer-tikka", "paneer shashlik"),
    ("ASC047", "daliya", "wheat daliya"),
    ("BFP177", "lemon-rasam", "nimbu rasam"),
    ("ASC190", "aloo-matar", "aloo mutter"),
    ("OSR112", "pav-bhaji", "pavbhaji"),
    ("ASC226", "kadhai-paneer", "karahi paneer"),
    ("ASC114", "plain-pulao", "zeera rice pulao"),
    ("BFP041", "rice-upma", "chawal upma"),
    ("ASC156", "whole-moong-dal", "green gram dal"),
    ("ASC167", "sambar", "south indian sambar"),
    ("ASC096", "roti", "chapati roti"),
    ("ASC126", "curd-rice", "dahi bhaat,thayir saadam"),
    ("ASC240", "chicken-curry", "chicken gravy"),
    ("ASC061", "omelette", "egg omelette"),
    ("BFP144", "plain-khichdi", "moong khichdi"),
    ("OSR104", "methi-thepla", "gujarati thepla"),
    ("ASC142", "naan", "butter naan"),
    ("ASC157", "masoor-dal", "red lentil dal"),
    ("ASC151", "moong-dal", "moong dal tadka"),
    ("ASC123", "vegetable-biryani", "veg dum biryani"),
    ("ASC177", "baingan-bharta", "eggplant bharta"),
    ("ASC273", "cucumber-raita", "dahi raita"),
    ("ASC146", "masala-dosa", "potato masala dosa"),
    ("BFP039", "rava-upma", "upma rava"),
    ("OSR139", "dal-makhani", "black dal"),
    ("ASC293", "suji-halwa", "sooji halwa"),
    ("ASC381", "paneer-tikka", "tandoori paneer"),
    ("ASC474", "dhokla", "gujarati dhokla"),
    ("BFP176", "rasam", "tomato rasam"),
    ("ASC059", "anda-bhurji", "bhurji"),
    ("ASC056", "boiled-egg", "anda uble"),
    ("ASC115", "mixed-vegetable-pulao", "veg pulav"),
    ("ASC098", "aloo-paratha", "aloo ka paratha"),
    ("ASC097", "paratha", "lachha paratha"),
    ("ASC195", "paneer-curry", "paneer masala"),
    ("ASC191", "matar-paneer", "mutter paneer"),
    ("ASC022", "salted-lassi", "salt lassi"),
    ("BFP240", "egg-curry", "ande ki curry"),
    ("OSR105", "rice-puttu", "puttu rice"),
    ("BFP153", "appam", "appam dosa"),
    ("ASC124", "lemon-rice", "chitrannam"),
    ("ASC127", "tamarind-rice", "pulihora tamarind"),
    ("ASC152", "urad-dal", "urad dal tadka"),
    ("ASC155", "mixed-dal", "panchmel dal"),
    ("ASC173", "cabbage-peas", "band gobi matar"),
    ("ASC269", "onion-tomato-raita", "pyaaz tamatar raita"),
    ("ASC147", "rava-dosa", "rava dose"),
    ("ASC150", "makki-roti", "makai roti"),
    ("ASC122", "mutton-biryani", "mutton dum biryani"),
    ("BFP043", "vegetable-upma", "upma vegetable"),
    ("BFP576", "poshtik-khichdi", "nutritious khichdi"),
    ("OSR010", "saffron-milk", "kesar milk"),
    ("OSR151", "jowar-dosa", "jowar dose"),
    ("ASC385", "peanut-chutney", "groundnut chutney"),
    ("ASC295", "gajar-halwa", "gajrela"),
    ("BFP040", "seviyan-upma", "semiya upma"),
    ("OSR099", "sabudana-khichdi", "sago khichri"),
    ("BFP439", "khaman", "yellow dhokla"),
    ("ASC047", "daliya", "broken wheat porridge"),
    ("ASC099", "mooli-paratha", "mooli ka paratha"),
    ("ASC105", "paneer-paratha", "cottage cheese paratha"),
    ("ASC472", "paushtik-roti", "fortified roti"),
    ("BFP177", "lemon-rasam", "nimbu rasam"),
)

DRINK_SLUGS = frozenset(
    {
        "chai",
        "lassi",
        "salted-lassi",
        "saffron-milk",
    }
)

DETECT_SERVING_UNIT = {
    "idli": "idli",
    "dosa": "dosa",
    "gulab-jamun": "piece",
    "samosa": "piece",
    "onion-pakoda": "piece",
    "rasmalai": "piece",
    "kulfi": "piece",
    "white-rice": "plate",
    "dal-tadka": "bowl",
    "palak-paneer": "bowl",
    "shahi-paneer": "bowl",
    "rajma-curry": "bowl",
    "chole": "bowl",
    "mutton-curry": "bowl",
    "fish-curry": "bowl",
    "chicken-biryani": "plate",
    "aloo-gobi": "bowl",
    "aloo-fry": "bowl",
    "dum-aloo": "bowl",
    "bhindi-masala": "bowl",
    "poha": "plate",
    "kheer": "bowl",
    "green-chutney": "tablespoon",
    "coconut-chutney": "tablespoon",
    "chicken-seekh-kebab": "piece",
}


def _q(value: Decimal) -> Decimal:
    return value.quantize(QUANTIZE, rounding=ROUND_HALF_UP)


def _cell(row: dict[str, str], key: str) -> str:
    return (row.get(key) or "").strip()


def _serving_grams(energy_kcal: Decimal, unit_kcal: Decimal | None) -> Decimal | None:
    if unit_kcal is None or energy_kcal <= 0:
        return None
    grams = (unit_kcal / energy_kcal) * Decimal("100")
    if grams < SERVING_MIN:
        return SERVING_MIN
    if grams > SERVING_MAX:
        return SERVING_MAX
    return _q(grams)


def _display_from_slug(slug: str) -> str:
    return slug.replace("-", " ")


def _load_indb(path: Path) -> dict[str, dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    by_code: dict[str, dict] = {}
    for row in rows[1:]:
        code = row[idx["food_code"]]
        if not code:
            continue
        energy = row[idx["energy_kcal"]]
        unit_kcal = row[idx["unit_serving_energy_kcal"]]
        by_code[str(code).strip()] = {
            "food_code": str(code).strip(),
            "food_name": str(row[idx["food_name"]] or "").strip(),
            "energy_kcal": energy,
            "protein_g": row[idx["protein_g"]],
            "carb_g": row[idx["carb_g"]],
            "fat_g": row[idx["fat_g"]],
            "servings_unit": str(row[idx["servings_unit"]] or "").strip(),
            "unit_serving_energy_kcal": unit_kcal,
        }
    return by_code


def _parse_source_code(source: str) -> str:
    parts = source.split()
    if len(parts) >= 2 and parts[0] == "INDB":
        return parts[1]
    return ""


def _aliases_for_detect(label: str) -> str:
    spaced = label.replace("-", " ")
    extras = {
        "aloo-gobi": "aloo gobhi,aloo gobi,potato cauliflower",
        "white-rice": "rice,chawal,boiled rice,steamed rice",
        "dal-tadka": "dal tadka,tadka dal,moong tadka",
        "chicken-biryani": "biryani,chicken pulao",
        "chai": "tea,garam chai,hot tea",
        "lassi": "sweet lassi,meethi lassi",
        "idli": "idly",
        "dosa": "dose,plain dosa",
        "chole": "chole masala,chana masala,channa curry",
        "gulab-jamun": "gulab jamun",
        "onion-pakoda": "onion pakora,pyaaz pakoda",
        "green-chutney": "hari chutney,mint coriander chutney",
        "coconut-chutney": "nariyal chutney",
    }
    return ",".join(filter(None, [spaced, extras.get(label, "")]))


def main() -> int:
    if not INDB_XLSX.is_file():
        print(f"missing INDB workbook: {INDB_XLSX}", file=sys.stderr)
        return 1
    if not DISH_CSV.is_file():
        print(f"missing {DISH_CSV}", file=sys.stderr)
        return 1

    indb = _load_indb(INDB_XLSX)
    used_codes: set[str] = set()
    out_rows: list[dict[str, str]] = []

    with DISH_CSV.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            class_id = int(_cell(row, "class_id"))
            label = _cell(row, "label")
            if FOOD_CLASS_LABELS[class_id] != label:
                print(f"label mismatch class_id={class_id}", file=sys.stderr)
                return 1
            source = _cell(row, "source")
            source_id = _parse_source_code(source)
            if source_id:
                used_codes.add(source_id)
            numeric = all(
                _cell(row, key)
                for key in (
                    "calories_per_100g",
                    "protein_per_100g",
                    "carb_per_100g",
                    "fat_per_100g",
                    "default_serving_grams",
                )
            )
            density = "1.00" if label in DRINK_SLUGS and numeric else ""
            serving_unit = DETECT_SERVING_UNIT.get(label, "")
            if label in DRINK_SLUGS:
                serving_unit = "tall glass"
            out_rows.append(
                {
                    "slug": label,
                    "name": _display_from_slug(label),
                    "detect_class_id": str(class_id),
                    "status": "complete" if numeric else "incomplete",
                    "calories_per_100g": _cell(row, "calories_per_100g"),
                    "protein_per_100g": _cell(row, "protein_per_100g"),
                    "carb_per_100g": _cell(row, "carb_per_100g"),
                    "fat_per_100g": _cell(row, "fat_per_100g"),
                    "density_g_per_ml": density,
                    "source_dataset": "indb" if source else "",
                    "source_id": source_id,
                    "source_note": source,
                    "aliases": _aliases_for_detect(label),
                    "indb_serving_unit": serving_unit,
                    "default_serving_grams": _cell(row, "default_serving_grams"),
                }
            )

    seen_slugs = {row["slug"] for row in out_rows}
    skipped_fat = []
    skipped_missing = []
    for code, slug, aliases in STAPLES:
        if slug in seen_slugs:
            continue
        recipe = indb.get(code)
        if recipe is None:
            skipped_missing.append((code, slug))
            continue
        if code in used_codes:
            continue
        energy = recipe["energy_kcal"]
        fat = recipe["fat_g"]
        if energy is None or fat is None:
            skipped_missing.append((code, slug))
            continue
        energy_d = Decimal(str(energy))
        fat_d = Decimal(str(fat))
        if fat_d > MAX_FAT_PER_100G:
            skipped_fat.append((code, slug, fat_d))
            continue
        unit_kcal = recipe["unit_serving_energy_kcal"]
        unit_d = Decimal(str(unit_kcal)) if unit_kcal is not None else None
        serving = _serving_grams(energy_d, unit_d)
        if serving is None:
            skipped_missing.append((code, slug))
            continue
        used_codes.add(code)
        seen_slugs.add(slug)
        density = "1.00" if slug in DRINK_SLUGS else ""
        name = recipe["food_name"] or _display_from_slug(slug)
        out_rows.append(
            {
                "slug": slug,
                "name": name,
                "detect_class_id": "",
                "status": "complete",
                "calories_per_100g": str(_q(energy_d)),
                "protein_per_100g": str(_q(Decimal(str(recipe["protein_g"])))),
                "carb_per_100g": str(_q(Decimal(str(recipe["carb_g"])))),
                "fat_per_100g": str(_q(fat_d)),
                "density_g_per_ml": density,
                "source_dataset": "indb",
                "source_id": code,
                "source_note": f"INDB {code} {name}",
                "aliases": aliases,
                "indb_serving_unit": recipe["servings_unit"],
                "default_serving_grams": str(serving),
            }
        )

    fieldnames = [
        "slug",
        "name",
        "detect_class_id",
        "status",
        "calories_per_100g",
        "protein_per_100g",
        "carb_per_100g",
        "fat_per_100g",
        "density_g_per_ml",
        "source_dataset",
        "source_id",
        "source_note",
        "aliases",
        "indb_serving_unit",
        "default_serving_grams",
    ]
    with OUT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    complete = sum(1 for row in out_rows if row["status"] == "complete")
    incomplete = sum(1 for row in out_rows if row["status"] == "incomplete")
    print(f"wrote {OUT_CSV}")
    print(f"foods={len(out_rows)} complete={complete} incomplete={incomplete}")
    if skipped_fat:
        print("skipped high-fat:", skipped_fat)
    if skipped_missing:
        print("skipped missing:", skipped_missing)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
